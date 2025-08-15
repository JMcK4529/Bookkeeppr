import calendar
import logging
import webview
from collections import defaultdict
from datetime import datetime, time
from io import BytesIO
from flask import (
    flash,
    make_response,
    redirect,
    render_template,
    url_for,
    request,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
from sqlite3 import IntegrityError
from webview import FileDialog
from lib.db import utils as dbutils
from lib.db.purchase import Purchase, PurchaseRepository
from lib.db.sale import Sale, SaleRepository

logger = logging.getLogger(__name__)


class SchedulerConfig:
    """
    Configuration class for APScheduler
    """

    SCHEDULER_API_ENABLED = True
    JOBS = [
        {
            "id": "daily_cleanup",
            "func": "lib.db.utils:delete_old_recovery_dbs",
            "args": (30,),  # older_than_days
            "trigger": "cron",
            "hour": 12,
            "minute": 0,
        }
    ]


def create_single_entity_view(entity_name, model_class, repo_class):
    def view(**kwargs):
        entity_id = kwargs[f"{entity_name[:-1]}_id"]
        repo = repo_class()
        match request.method:
            case "GET":
                obj = repo.read(id=entity_id)
                template = f"{entity_name[:-1]}.html"
                context_key = f"{entity_name[:-1]}_obj"

                if obj:
                    return render_template(template, **{context_key: obj})
                else:
                    flash(
                        f"{entity_name[:-1].capitalize()} does not exist",
                        "error",
                    )
                    return redirect(
                        request.referrer or url_for(f"{entity_name}")
                    )
            case "PATCH":
                name = request.form.get("name", "").strip()
                try:
                    repo.update(model_class(id=entity_id, name=name))
                    flash(
                        f"{entity_name[:-1].capitalize()} updated successfully",
                        "success",
                    )
                    response = make_response("OK", 204)
                except IntegrityError as err:
                    if "UNIQUE constraint failed" in str(err):
                        err_msg = f"{name} already exists"
                    else:
                        err_msg = f"Unexpected error: {str(err)}"
                    logger.warning(err_msg)
                    flash(err_msg, "error")
                    response = make_response(f"Update failed. {err_msg}", 409)
                response.headers["Content-Type"] = "text/plain; charset=utf-8"
                return response
            case "DELETE":
                # Assume by this point we have safely confirmed the delete command
                entity = repo.read(id=entity_id)
                if not entity:
                    err_msg = f"{entity_name[:-1].capitalize()} not found."
                    logger.warning(f"[DELETE] {err_msg}")
                    flash(err_msg, "error")
                    response = make_response(err_msg, 404)
                    response.headers["Content-Type"] = (
                        "text/plain; charset=utf-8"
                    )
                    return response

                safe_to_delete = False
                # Create the recovery DB
                try:
                    dbutils.backup_deleted_entity(entity, repo_class)
                    safe_to_delete = True
                except Exception as err:
                    err_msg = f"Data backup failed. Delete operation on {entity.name} aborted."
                    logger.warning(f"[DELETE] {err_msg}.\n" + f"{err}")
                    flash(err_msg, "error")
                    response = make_response(err_msg, 500)

                if safe_to_delete:
                    try:
                        repo.delete(entity.id)
                        flash(
                            f"{entity.name} was successfully deleted.",
                            "success",
                        )
                        response = make_response("OK", 204)
                    except IntegrityError as err:
                        err_msg = f"Delete blocked due to data constraint: {str(err)}"
                        logger.warning(f"[DELETE] {err_msg}")
                        flash(err_msg, "error")
                        response = make_response(err_msg, 409)
                    except Exception as err:
                        err_msg = f"Unexpected error: {str(err)}"
                        logger.error(f"[DELETE] {err_msg}")
                        flash(f"{entity.name} could not be deleted.", "error")
                        response = make_response(f"{err_msg}", 500)

                response.headers["Content-Type"] = "text/plain; charset=utf-8"
                return response

    return view


def register_entity_routes(
    entity_name, template_name, model_class, repo_class, app
):
    list_endpoint = f"/{entity_name}"
    create_endpoint = f"/{entity_name}/create"
    by_id_endpoint = f"/{entity_name}/<int:{entity_name[:-1]}_id>"

    list_endpoint_name = f"{entity_name}"
    create_endpoint_name = f"create_{entity_name}"
    by_id_endpoint_name = f"single_{entity_name[:-1]}"

    @app.route(list_endpoint, endpoint=list_endpoint_name, methods=["GET"])
    def list_entities():
        query = request.args.get("q", "")
        repo = repo_class()
        results = repo.search(query) if query else repo.all()
        return render_template(
            template_name, query=query, **{entity_name: results}
        )

    @app.route(
        create_endpoint, methods=["POST"], endpoint=create_endpoint_name
    )
    def create_entity():
        name = request.form.get("name")
        repo = repo_class()
        try:
            new_entity = repo.create(model_class(None, name))
            flash(f"Successfully added {entity_name[:-1]}: {name}", "success")
        except IntegrityError as err:
            if "UNIQUE constraint failed" in str(err):
                err_msg = f"{name} already exists"
            else:
                err_msg = f"Unexpected Error: {str(err)}"
            logger.warning(err_msg)
            flash(err_msg, "error")
        return redirect(list_endpoint)

    app.add_url_rule(
        rule=by_id_endpoint,
        endpoint=by_id_endpoint_name,
        view_func=create_single_entity_view(
            entity_name, model_class, repo_class
        ),
        methods=["GET", "PATCH", "DELETE"],
    )


###
def build_filters(model_class, request):
    def parse_range(prefix):
        result = {}
        min_val = request.args.get(f"{prefix}_min")
        max_val = request.args.get(f"{prefix}_max")
        eq_val = request.args.get(f"{prefix}_eq")
        if eq_val:
            result["eq"] = float(eq_val)
        else:
            if min_val:
                result["min"] = float(min_val)
            if max_val:
                result["max"] = float(max_val)
        return result

    filters = {
        "net": parse_range("net"),
        "vat": [],
        "payment": [],
        "timeFrom": request.args.get("timeFrom", "").strip(),
        "timeTo": request.args.get("timeTo", "").strip(),
    }
    match model_class.__name__:
        case "Sale":
            filters.update(
                {
                    "customer": request.args.get("customer", "").strip(),
                    "invoice": request.args.get("invoice", "").strip(),
                }
            )
        case "Purchase":
            filters.update(
                {
                    "supplier": request.args.get("supplier", "").strip(),
                    "supplier_invoice": request.args.get(
                        "supplier_invoice", ""
                    ).strip(),
                    "internal_invoice": request.args.get(
                        "internal_invoice", ""
                    ).strip(),
                    "goods": parse_range("goods"),
                    "utilities": parse_range("utilities"),
                    "motor_expenses": parse_range("motor_expenses"),
                    "sundries": parse_range("sundries"),
                    "miscellaneous": parse_range("miscellaneous"),
                    "capital_spend": request.args.get(
                        "capital_spend", ""
                    ).strip(),
                }
            )
    vat_values = request.args.getlist("vat")
    payment_values = request.args.getlist("payment")
    if vat_values:
        filters.update({"vat": [float(v) for v in vat_values if v]})
    if payment_values:
        filters.update({"payment": [p for p in payment_values if p]})
    return filters


def extract_transaction_to_create(
    form, model_class, entity_class_name, entities
):
    entity_name = form.get(f"{entity_class_name.lower()}_name")
    entity_obj = next((e for e in entities if e.name == entity_name), None)
    if not entity_obj:
        raise ValueError(f"Invalid {entity_class_name.lower()} selected")

    args_dict = {
        "id": None,
        f"{entity_class_name.lower()}_id": entity_obj.id,
        f"{entity_class_name.lower()}_name": entity_obj.name,
        "net_amount": float(request.form["net_amount"]),
        "vat_percent": float(request.form["vat_percent"]),
        "payment_method": request.form["payment_method"],
        "timestamp": request.form["timestamp"],
    }

    match model_class.__name__:
        case "Sale":
            args_dict.update(
                {"invoice_number": request.form.get("invoice_number")}
            )
        case "Purchase":
            args_dict.update(
                {
                    "supplier_invoice_code": request.form.get(
                        "supplier_invoice_code"
                    ),
                    "internal_invoice_number": request.form.get(
                        "internal_invoice_number"
                    ),
                    "goods": float(request.form.get("goods")),
                    "utilities": float(request.form.get("utilities")),
                    "motor_expenses": float(
                        request.form.get("motor_expenses")
                    ),
                    "sundries": float(request.form.get("sundries")),
                    "miscellaneous": float(request.form.get("miscellaneous")),
                    "capital_spend": bool(request.form.get("capital_spend")),
                }
            )

    new_transaction = model_class(**args_dict)
    return new_transaction


def create_single_transaction_view(
    transaction_name,
    model_class,
    repo_class,
    entity_class_name,
    entity_repo_class,
):
    def view(**kwargs):
        entity_repo = entity_repo_class()
        entities = entity_repo.all()
        transaction_id = kwargs[f"{transaction_name[:-1]}_id"]
        repo = repo_class()
        match request.method:
            case "GET":
                obj = repo.read(id=transaction_id)
                template = f"{transaction_name[:-1]}.html"
                context_key = f"{transaction_name[:-1]}_obj"

                if obj:
                    return render_template(
                        template,
                        **{
                            context_key: obj,
                            f"{entity_class_name.lower()}s": entities,
                        },
                    )
                else:
                    flash(
                        f"{transaction_name[:-1].capitalize()} does not exist",
                        "error",
                    )
                    return redirect(
                        request.referrer or url_for(f"{transaction_name}")
                    )
            case "PATCH":
                obj = repo.read(id=transaction_id)
                if not obj:
                    err_msg = (
                        f"{transaction_name[:-1].capitalize()} not found."
                    )
                    logger.warning(f"[PATCH] {err_msg}")
                    flash(err_msg, "error")
                    return make_response(err_msg, 404)

                try:
                    updates = {}
                    for key in request.form:
                        val = request.form[key]
                        if key == "vat_percent":
                            updates[key] = float(
                                val
                            )  # Already a decimal from JS
                        elif key in {
                            "net_amount",
                            "goods",
                            "utilities",
                            "motor_expenses",
                            "sundries",
                            "miscellaneous",
                        }:
                            updates[key] = float(val)
                        elif key in {"capital_spend"}:
                            updates[key] = bool(val.lower() == "true")
                        else:
                            updates[key] = val

                    for attr, val in updates.items():
                        setattr(obj, attr, val)

                    repo.update(obj)
                    flash(
                        f"{transaction_name[:-1].capitalize()} updated successfully.",
                        "success",
                    )
                    return make_response("OK", 204)
                except Exception as err:
                    err_msg = f"Update failed: {str(err)}"
                    logger.error(f"[PATCH] {err_msg}")
                    flash(err_msg, "error")
                    return make_response(err_msg, 500)
            case "DELETE":
                transaction = repo.read(id=transaction_id)
                if not transaction:
                    err_msg = (
                        f"{transaction_name[:-1].capitalize()} not found."
                    )
                    logger.warning(f"[DELETE] {err_msg}")
                    flash(err_msg, "error")
                    response = make_response(err_msg, 404)
                    response.headers["Content-Type"] = (
                        "text/plain; charset=utf-8"
                    )
                    return response

                transaction_invoice = None
                match model_class.__name__:
                    case "Sale":
                        transaction_invoice = transaction.invoice_number
                    case "Purchase":
                        transaction_invoice = (
                            transaction.internal_invoice_number
                        )

                safe_to_delete = False
                try:
                    dbutils.backup_deleted_transactions(
                        [transaction], repo_class
                    )
                    safe_to_delete = True
                except Exception as err:
                    err_msg = f"Data backup failed. Delete operation on {transaction_invoice} aborted."
                    logger.warning(f"[DELETE] {err_msg}.\n" + f"{err}")
                    flash(err_msg, "error")
                    response = make_response(err_msg, 500)

                if safe_to_delete:
                    try:
                        repo.delete(transaction.id)
                        flash(
                            f"{transaction_invoice} was successfully deleted.",
                            "success",
                        )
                        response = make_response("OK", 204)
                    except IntegrityError as err:
                        err_msg = f"Delete blocked due to data constraint: {str(err)}"
                        logger.warning(f"[DELETE] {err_msg}")
                        flash(err_msg, "error")
                        response = make_response(err_msg, 409)
                    except Exception as err:
                        err_msg = f"Unexpected error: {str(err)}"
                        logger.error(f"[DELETE] {err_msg}")
                        flash(
                            f"{transaction_invoice} could not be deleted.",
                            "error",
                        )
                        response = make_response(f"{err_msg}", 500)

                response.headers["Content-Type"] = "text/plain; charset=utf-8"
                return response

    return view


def register_transaction_routes(
    transaction_name,
    template_name,
    model_class,
    repo_class,
    entity_repo_class,
    entity_class_name,
    app,
):
    list_endpoint = f"/{transaction_name}"
    create_endpoint = f"/{transaction_name}/create"
    by_id_endpoint = f"/{transaction_name}/<int:{transaction_name[:-1]}_id>"

    list_endpoint_name = f"{transaction_name}"
    create_endpoint_name = f"create_{transaction_name}"
    by_id_endpoint_name = f"single_{transaction_name[:-1]}"

    repo = repo_class()

    @app.route(list_endpoint, endpoint=list_endpoint_name, methods=["GET"])
    def list_transactions():
        filters = build_filters(model_class, request)
        all_transactions = repo.all()

        vat_options = sorted(set(t.vat_percent for t in all_transactions))
        payment_options = sorted(
            set(t.payment_method for t in all_transactions)
        )
        transactions = (
            repo.search(filters) if any(filters.values()) else all_transactions
        )

        return render_template(
            template_name,
            filters=filters,
            **{transaction_name: transactions},
            vat_options=vat_options,
            payment_options=payment_options,
        )

    @app.route(
        create_endpoint, endpoint=create_endpoint_name, methods=["GET", "POST"]
    )
    def create_transaction():
        entity_repo = entity_repo_class()
        entities = entity_repo.all()
        render = render_template(
            f"create_{transaction_name[:-1]}.html",
            **{f"{entity_class_name.lower()}s": entities},
        )

        if request.method == "POST":
            try:
                new_obj = extract_transaction_to_create(
                    request.form, model_class, entity_class_name, entities
                )
                repo.create(new_obj)
                flash(
                    f"{transaction_name[:-1].capitalize()} created successfully.",
                    "success",
                )
                return redirect(create_endpoint)
            except ValueError as err:
                flash(f"{str(err)}", "error")
                return redirect(list_endpoint)
            except IntegrityError as err:
                logger.warning(err)
                flash(f"Unexpected Error: {str(err)}", "error")

        return render

    app.add_url_rule(
        rule=by_id_endpoint,
        endpoint=by_id_endpoint_name,
        view_func=create_single_transaction_view(
            transaction_name,
            model_class,
            repo_class,
            entity_class_name,
            entity_repo_class,
        ),
        methods=["GET", "PATCH", "DELETE"],
    )


###
# def export_to_xlsx(
#     transaction_type: str, start_date: str, end_date: str
# ) -> BytesIO:
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Summary"
#     ws.append(["Transaction Type", "Start Date", "End Date"])
#     ws.append([transaction_type, start_date, end_date])

#     file_stream = BytesIO()
#     wb.save(file_stream)
#     file_stream.seek(0)
#     return file_stream


def export_to_xlsx(
    transaction_name: str, start_date: str, end_date: str, file: Path | BytesIO
) -> bool | BytesIO:
    """
    Creates a spreadsheet including transactions of type `transaction_name` which
    are timestamped between `start_date` and `end_date` (inclusive).
    If `file` is a Path, return `file.exists()`.
    If `file` is a BytesIO stream, return the stream.
    :param str transaction_name: Sales or Purchases
    :param str start_date: Start of date range for data to export
    :param str end_date: End of date range for data to export
    :param Path|BytesIO file: Destination filepath or filestream
    """
    repo = None
    match transaction_name.lower():
        case "sales":
            repo = SaleRepository()
        case "purchases":
            repo = PurchaseRepository()

    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    start_dt = datetime.combine(start_dt.date(), time.min)
    end_dt = datetime.combine(end_dt.date(), time.max)
    start_str = start_dt.strftime("%Y-%m-%d %H:%M:%S")
    end_str = end_dt.replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")

    # Get all a list of all <transaction_name>s from the date range, sorted by timestamp
    transactions = repo.search(
        {
            "timeFrom": start_str,
            "timeTo": end_str,
        }
    )

    # Create the front sheet
    wb = Workbook()
    ws_front = wb.active
    ws_front.title = "Front Sheet"
    ws_front.append(["Transaction Type", "Start Date", "End Date"])
    ws_front.append([transaction_name, start_date, end_date])
    ws_front.append([])

    # Group transactions by (year, month)
    months = defaultdict(list)
    for t in transactions:
        ts = datetime.strptime(t.timestamp, "%Y-%m-%d %H:%M:%S")
        key = (ts.year, ts.month)
        months[key].append((ts, t))

    for (year, month), entries in sorted(months.items()):
        sheet_name = f"{calendar.month_abbr[month]} '{str(year)[-2:]}"
        ws = wb.create_sheet(title=sheet_name)

        if transaction_name.lower() == "sales":
            ws.append(
                [
                    "Customer",
                    "Invoice Number",
                    "Net",
                    "VAT%",
                    "VAT",
                    "Total",
                    "Payment Method",
                    "Date",
                ]
            )
            for i, (ts, t) in enumerate(entries, start=2):
                row = [
                    t.customer_name,
                    t.invoice_number,
                    t.net_amount,
                    t.vat_percent,
                    f"=C{i}*D{i}",
                    f"=C{i}+E{i}",
                    t.payment_method,
                    ts.strftime("%d/%m/%Y"),
                ]
                ws.append(row)
            ws.append(
                [
                    "TOTAL",
                    "",
                    f"=SUM(C2:C{len(entries)+1})",
                    "",
                    f"=SUM(E2:E{len(entries)+1})",
                    f"=SUM(F2:F{len(entries)+1})",
                    "",
                    "",
                ]
            )
        else:
            ws.append(
                [
                    "Supplier",
                    "Invoice Number",
                    "Net",
                    "Goods",
                    "Utilities",
                    "Motor Expenses",
                    "Sundries",
                    "Miscellaneous",
                    "VAT%",
                    "VAT",
                    "Total",
                    "Supplier Invoice",
                    "Payment Method",
                    "Date",
                ]
            )
            for i, (ts, t) in enumerate(entries, start=2):
                row = [
                    t.supplier_name,
                    t.internal_invoice_number,
                    t.net_amount,
                    t.goods,
                    t.utilities,
                    t.motor_expenses,
                    t.sundries,
                    t.miscellaneous,
                    t.vat_percent,
                    f"=C{i}*I{i}",
                    f"=C{i}+J{i}",
                    t.supplier_invoice_code,
                    t.payment_method,
                    ts.strftime("%d/%m/%Y"),
                ]
                ws.append(row)
            ws.append(
                [
                    "TOTAL",
                    "",
                    f"=SUM(C2:C{len(entries)+1})",
                    f"=SUM(D2:D{len(entries)+1})",
                    f"=SUM(E2:E{len(entries)+1})",
                    f"=SUM(F2:F{len(entries)+1})",
                    f"=SUM(G2:G{len(entries)+1})",
                    f"=SUM(H2:H{len(entries)+1})",
                    "",
                    f"=SUM(J2:J{len(entries)+1})",
                    f"=SUM(K2:K{len(entries)+1})",
                    "",
                    "",
                    "",
                ]
            )

        # Style headers
        header_fill = PatternFill(
            start_color="5E1791", end_color="5E1791", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Set number formats
        if transaction_name.lower() == "sales":
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    cell.number_format = "£#,##0.00"
            for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
                for cell in row:
                    cell.number_format = "0%;-0%;0%"
            for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
                for cell in row:
                    cell.number_format = "£#,##0.00"
        else:
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=8):
                for cell in row:
                    cell.number_format = "£#,##0.00"
            for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
                for cell in row:
                    cell.number_format = "0%;-0%;0%"
            for row in ws.iter_rows(min_row=2, min_col=10, max_col=11):
                for cell in row:
                    cell.number_format = "£#,##0.00"

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

    wb.save(file)
    if isinstance(file, BytesIO):
        file.seek(0)
        return file
    else:
        return file.exists()


def open_export_file_picker(transaction_name: str) -> Path:
    default_filename = f"{transaction_name.capitalize()} Record.xlsx"

    window = webview.windows[0]
    selected = window.create_file_dialog(
        dialog_type=FileDialog.SAVE,
        file_types=("Excel files (*.xlsx)",),
        save_filename=default_filename,
    )

    if selected:
        if isinstance(selected, (list, tuple)):
            file_path = selected[0]
        else:
            file_path = selected

        if isinstance(file_path, str):
            path = Path(file_path)
        elif isinstance(file_path, Path):
            path = file_path
        else:
            path = Path(str(file_path))
        logger.info(f"[EXPORT] Final save path: {path}")
        return path

    return None
